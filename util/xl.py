import xlrd
from openpyxl import load_workbook


class read_xl:
    def read_xls_dict(self, row_id=1, dest="", data=""):
        # 打开工作薄
        workbook = xlrd.open_workbook(dest)
        # 获取第一个sheet表格
        table = workbook.sheets()[0]
        # 获取行数
        rows = table.nrows
        # 获取列数
        cols = table.ncols

        dic = {}
        # 循环获取每行的数据
        name = ""
        for row in range(row_id, rows):
            for col in range(0, cols):
                value = table.cell_value(row, col)
                if col == 0:
                    name = value
                elif col == 1:
                    dic[name] = str(value) + data
        return dic

    def read_xlsx_dict(self, row_id=1, dest="", data=""):
        # openpyxl 是从1开始的
        row_id += 1
        workbook = load_workbook(dest)
        sheet = workbook["Sheet1"]
        dic = {}
        name = ""
        for row in range(row_id, sheet.max_row + 1):
            for column in range(1, sheet.max_column + 1):
                value = sheet.cell(row=row, column=column).value
                if column == 1:
                    name = value
                elif column == 2:
                    dic[name] = str(value) + data

        return dic

# if __name__ == '__main__':
# xl = read_xl()
# list_dict = xl.get_list_dict(dest=r"../xl/java.xls",row_id=1)
# for name in list_dict.keys():
#     print(name, ",", list_dict[name])
